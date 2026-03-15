# Create an improved habit tracker spreadsheet with progress calculations and a simple chart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

wb = Workbook()

# Sheet 1: Habit Tracker
ws = wb.active
ws.title = "Tracker"

ws["A1"] = "Monthly Habit Tracker"
ws["A1"].font = Font(size=16, bold=True)
ws.merge_cells("A1:AI1")
ws["A1"].alignment = Alignment(horizontal="center")

# Header
ws["A3"] = "Habit"
for day in range(1, 32):
    ws.cell(row=3, column=day+1).value = day

ws.cell(row=3, column=33).value = "Total"
ws.cell(row=3, column=34).value = "Completion %"

# Habit rows
for i in range(10):
    row = 4 + i
    ws.cell(row=row, column=1).value = f"Habit {i+1}"
    ws.cell(row=row, column=33).value = f'=COUNTIF(B{row}:AF{row},"✔")'
    ws.cell(row=row, column=34).value = f"=AG{row}/31"

# Column widths
ws.column_dimensions["A"].width = 22
for col in range(2, 35):
    ws.column_dimensions[get_column_letter(col)].width = 4

# Sheet 2: Dashboard
dash = wb.create_sheet("Dashboard")
dash["A1"] = "Habit Progress Dashboard"
dash["A1"].font = Font(size=16, bold=True)

dash["A3"] = "Habit"
dash["B3"] = "Completion %"

for i in range(10):
    dash.cell(row=4+i, column=1).value = f"=Tracker!A{4+i}"
    dash.cell(row=4+i, column=2).value = f"=Tracker!AH{4+i}"

# Create chart
chart = BarChart()
chart.title = "Habit Completion"
data = Reference(dash, min_col=2, min_row=3, max_row=13)
cats = Reference(dash, min_col=1, min_row=4, max_row=13)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
dash.add_chart(chart, "D3")

file_path = "/mnt/data/advanced_habit_tracker.xlsx"
wb.save(file_path)

file_path